import logging
import os
import re
import tempfile

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.http import url_has_allowed_host_and_scheme
from django.views import View

import json

from .filtros import (
    FECHA_OPCIONES, OPERADORES_POR_TIPO,
    apply_filters, parse_filtros_from_post,
)
from .importar import auto_detect_mapping, detect_tipo, import_from_rows, parse_file
from .models import CampoPersonalizado, Contacto, ValorCampo

logger = logging.getLogger('apps.whatsapp')

TIPOS_CAMPO = CampoPersonalizado.TIPOS


def _grupos_existentes():
    g1 = set(CampoPersonalizado.objects.values_list('grupo', flat=True).exclude(grupo=''))
    g2 = set(Contacto.objects.values_list('grupo', flat=True).exclude(grupo=''))
    return sorted(g1 | g2)


def _campos_para_grupo(grupo=''):
    return CampoPersonalizado.objects.filter(activo=True).filter(
        Q(grupo='') | Q(grupo=grupo)
    ).order_by('orden', 'etiqueta')


# ──────────────────────────────────────────────
# Contactos
# ──────────────────────────────────────────────

class ContactoListView(LoginRequiredMixin, View):
    def get(self, request):
        qs = Contacto.objects.all()
        q = request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(Q(nombre__icontains=q) | Q(telefono__icontains=q) | Q(email__icontains=q))
        grupo = request.GET.get('grupo', '').strip()
        if grupo:
            qs = qs.filter(grupo=grupo)

        grupos = Contacto.objects.values_list('grupo', flat=True).distinct().exclude(grupo='').order_by('grupo')
        total = qs.count()
        page = max(1, int(request.GET.get('p', 1) or 1))
        per_page = 50
        offset = (page - 1) * per_page
        contactos = list(qs[offset:offset + per_page])

        return render(request, 'contacts/list.html', {
            'contactos': contactos, 'q': q, 'grupo': grupo, 'grupos': grupos,
            'total': total, 'page': page,
            'has_prev': page > 1, 'has_next': offset + per_page < total,
            'prev_page': page - 1, 'next_page': page + 1,
        })


class ContactoExportarView(LoginRequiredMixin, View):
    def get(self, request):
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        qs = Contacto.objects.all().order_by('nombre')
        q = request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(Q(nombre__icontains=q) | Q(telefono__icontains=q) | Q(email__icontains=q))
        grupo = request.GET.get('grupo', '').strip()
        if grupo:
            qs = qs.filter(grupo=grupo)

        campos = list(CampoPersonalizado.objects.filter(activo=True).order_by('orden', 'etiqueta'))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Contactos'

        headers = ['Nombre', 'Teléfono', 'Email', 'Grupo', 'Notas', 'Fecha de creación'] + [c.etiqueta for c in campos]
        ws.append(headers)

        for c in qs.prefetch_related('valores'):
            val_map = {v.campo_id: v.valor for v in c.valores.all()}
            ws.append([
                c.nombre, c.telefono, c.email, c.grupo, c.notas,
                c.created_at.strftime('%d/%m/%Y %H:%M') if c.created_at else '',
            ] + [val_map.get(campo.pk, '') for campo in campos])

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 22

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="contactos.xlsx"'
        wb.save(response)
        return response


class ContactoDetailView(LoginRequiredMixin, View):
    def get(self, request, pk):
        from apps.whatsapp.models import Mensaje as MensajeModel
        contacto = get_object_or_404(Contacto, pk=pk)
        campos_con_valores = contacto.get_campos_con_valores()
        conversacion = None
        try:
            conversacion = contacto.conversaciones.order_by('-ultimo_mensaje_at').first()
        except Exception:
            pass
        archivos = []
        try:
            conv_ids = list(contacto.conversaciones.values_list('pk', flat=True))
            if conv_ids:
                qs = MensajeModel.objects.filter(
                    conversacion_id__in=conv_ids,
                    tipo__in=['image', 'document', 'audio', 'video']
                ).order_by('-timestamp')[:50]
                # Campos opcionales — pueden no existir si la migración no corrió
                try:
                    archivos = list(qs.values(
                        'tipo', 'media_url', 'media_filename', 'media_mime', 'timestamp', 'direccion'
                    ))
                except Exception:
                    archivos = list(qs.values('tipo', 'media_url', 'timestamp', 'direccion'))
                    for a in archivos:
                        a.setdefault('media_filename', '')
                        a.setdefault('media_mime', '')
        except Exception:
            archivos = []
        return render(request, 'contacts/detail.html', {
            'contacto': contacto,
            'campos_con_valores': campos_con_valores,
            'conversacion': conversacion,
            'archivos': archivos,
        })


class ContactoCreateView(LoginRequiredMixin, View):
    template_name = 'contacts/form.html'

    def get(self, request):
        grupo = request.GET.get('grupo', '')
        campos = list(_campos_para_grupo(grupo))
        return render(request, self.template_name, {
            'campos': campos,
            'grupos': _grupos_existentes(),
            'tipos': TIPOS_CAMPO,
            'prefill_tel': request.GET.get('telefono', ''),
            'prefill_nombre': request.GET.get('nombre', ''),
            'prefill_grupo': grupo,
        })

    def post(self, request):
        nombre = request.POST.get('nombre', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        email = request.POST.get('email', '').strip()
        grupo = request.POST.get('grupo', '').strip()
        notas = request.POST.get('notas', '').strip()

        if not nombre or not telefono:
            messages.error(request, 'Nombre y teléfono son requeridos.')
            return redirect(request.get_full_path())

        if not telefono.startswith('+'):
            telefono = '+' + telefono

        if Contacto.objects.filter(telefono=telefono).exists():
            messages.error(request, f'Ya existe un contacto con el teléfono {telefono}.')
            return redirect(request.get_full_path())

        contacto = Contacto.objects.create(
            nombre=nombre, telefono=telefono, email=email, grupo=grupo, notas=notas,
        )

        for campo in _campos_para_grupo(grupo):
            valor = request.POST.get(f'campo_{campo.pk}', '').strip()
            if valor:
                ValorCampo.objects.create(contacto=contacto, campo=campo, valor=valor)

        _auto_link_conversacion(telefono, contacto, nombre)

        messages.success(request, f'Contacto {nombre} creado.')
        return redirect('contacts:detail', pk=contacto.pk)


class ContactoUpdateView(LoginRequiredMixin, View):
    template_name = 'contacts/form.html'

    def get(self, request, pk):
        contacto = get_object_or_404(Contacto, pk=pk)
        campos = list(_campos_para_grupo(contacto.grupo))
        val_map = {v.campo_id: v.valor for v in contacto.valores.all()}
        campos_con_valores = [(c, val_map.get(c.pk, '')) for c in campos]
        return render(request, self.template_name, {
            'obj': contacto,
            'campos_con_valores': campos_con_valores,
            'grupos': _grupos_existentes(),
            'tipos': TIPOS_CAMPO,
            'prefill_nombre': contacto.nombre,
            'prefill_tel': contacto.telefono,
            'prefill_grupo': contacto.grupo,
        })

    def post(self, request, pk):
        contacto = get_object_or_404(Contacto, pk=pk)
        nombre = request.POST.get('nombre', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        email = request.POST.get('email', '').strip()
        grupo = request.POST.get('grupo', '').strip()
        notas = request.POST.get('notas', '').strip()

        if not nombre or not telefono:
            messages.error(request, 'Nombre y teléfono son requeridos.')
            return redirect(request.get_full_path())

        if not telefono.startswith('+'):
            telefono = '+' + telefono

        dup = Contacto.objects.filter(telefono=telefono).exclude(pk=pk).first()
        if dup:
            messages.error(request, f'El teléfono ya está en uso por {dup.nombre}.')
            return redirect(request.get_full_path())

        contacto.nombre = nombre
        contacto.telefono = telefono
        contacto.email = email
        contacto.grupo = grupo
        contacto.notas = notas
        contacto.save()

        for campo in _campos_para_grupo(grupo):
            valor = request.POST.get(f'campo_{campo.pk}', '').strip()
            if valor:
                ValorCampo.objects.update_or_create(
                    contacto=contacto, campo=campo, defaults={'valor': valor},
                )
            else:
                ValorCampo.objects.filter(contacto=contacto, campo=campo).delete()

        # Sync conversacion nombre
        try:
            from apps.whatsapp.models import Conversacion
            Conversacion.objects.filter(contacto=contacto).update(nombre_contacto=nombre)
        except Exception:
            pass

        messages.success(request, 'Contacto actualizado.')
        return redirect('contacts:detail', pk=contacto.pk)


class ContactoDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        contacto = get_object_or_404(Contacto, pk=pk)
        nombre = contacto.nombre
        contacto.delete()
        messages.success(request, f'Contacto {nombre} eliminado.')
        return redirect('contacts:list')


# ──────────────────────────────────────────────
# Campos personalizados
# ──────────────────────────────────────────────

class CampoListView(LoginRequiredMixin, View):
    def get(self, request):
        campos = CampoPersonalizado.objects.all()
        return render(request, 'contacts/campos_list.html', {
            'campos': campos,
            'grupos': _grupos_existentes(),
            'tipos': TIPOS_CAMPO,
        })


class CampoCreateView(LoginRequiredMixin, View):
    template_name = 'contacts/campo_form.html'

    def get(self, request):
        return render(request, self.template_name, {
            'grupos': _grupos_existentes(), 'tipos': TIPOS_CAMPO,
        })

    def post(self, request):
        etiqueta = request.POST.get('etiqueta', '').strip()
        tipo = request.POST.get('tipo', 'text')
        grupo = request.POST.get('grupo', '').strip()
        orden = request.POST.get('orden', '0').strip()

        if not etiqueta:
            messages.error(request, 'La etiqueta es requerida.')
            return redirect(request.get_full_path())

        nombre = re.sub(r'[^\w]', '_', etiqueta.lower())[:100]
        base = nombre
        counter = 1
        while CampoPersonalizado.objects.filter(nombre=nombre).exists():
            nombre = f'{base}_{counter}'
            counter += 1

        CampoPersonalizado.objects.create(
            nombre=nombre, etiqueta=etiqueta, tipo=tipo, grupo=grupo,
            orden=int(orden) if orden.isdigit() else 0,
        )
        messages.success(request, f'Campo "{etiqueta}" creado.')

        next_url = request.POST.get('next', '')
        if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
            return redirect(next_url)
        return redirect('contacts:campos')


class CampoUpdateView(LoginRequiredMixin, View):
    template_name = 'contacts/campo_form.html'

    def get(self, request, pk):
        campo = get_object_or_404(CampoPersonalizado, pk=pk)
        return render(request, self.template_name, {
            'obj': campo, 'grupos': _grupos_existentes(), 'tipos': TIPOS_CAMPO,
        })

    def post(self, request, pk):
        campo = get_object_or_404(CampoPersonalizado, pk=pk)
        campo.etiqueta = request.POST.get('etiqueta', campo.etiqueta).strip()
        campo.tipo = request.POST.get('tipo', campo.tipo)
        campo.grupo = request.POST.get('grupo', '').strip()
        orden = request.POST.get('orden', str(campo.orden)).strip()
        campo.orden = int(orden) if orden.isdigit() else campo.orden
        campo.activo = request.POST.get('activo') == 'on'
        campo.save()
        messages.success(request, 'Campo actualizado.')
        return redirect('contacts:campos')


class CampoDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        get_object_or_404(CampoPersonalizado, pk=pk).delete()
        messages.success(request, 'Campo eliminado.')
        return redirect('contacts:campos')


# ──────────────────────────────────────────────
# Importar contactos
# ──────────────────────────────────────────────

class ImportarContactosView(LoginRequiredMixin, View):
    template_name = 'contacts/importar.html'

    def get(self, request):
        tmp = request.session.pop('import_tmp', None)
        if tmp and os.path.exists(tmp):
            try:
                os.unlink(tmp)
            except Exception:
                pass
        return render(request, self.template_name)

    def post(self, request):
        step = request.POST.get('step', '1')
        if step == '1':
            return self._handle_upload(request)
        if step == '2':
            return self._handle_confirm(request)
        return redirect(request.path)

    def _handle_upload(self, request):
        file = request.FILES.get('archivo')
        if not file:
            messages.error(request, 'Seleccioná un archivo.')
            return redirect(request.path)

        ext = file.name.rsplit('.', 1)[-1].lower()
        if ext not in ('xlsx', 'xls', 'csv'):
            messages.error(request, 'Solo se permiten archivos .xlsx, .xls o .csv')
            return redirect(request.path)

        with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}', prefix='import_') as f:
            for chunk in file.chunks():
                f.write(chunk)
            tmp_path = f.name

        with open(tmp_path, 'rb') as f:
            headers, rows, error = parse_file(f, ext)

        if error:
            os.unlink(tmp_path)
            messages.error(request, error)
            return redirect(request.path)

        if not rows:
            os.unlink(tmp_path)
            messages.error(request, 'El archivo no tiene datos.')
            return redirect(request.path)

        auto_mapping = auto_detect_mapping(headers)
        col_tipos = {}
        for i, h in enumerate(headers):
            if auto_mapping.get(i) == 'campo':
                sample = [row[i] for row in rows[:50] if i < len(row)]
                col_tipos[i] = detect_tipo(sample)

        request.session['import_tmp'] = tmp_path
        request.session['import_ext'] = ext
        request.session['import_headers'] = headers
        request.session['import_total'] = len(rows)

        header_data = [
            {
                'i': i,
                'header': h,
                'auto_role': auto_mapping.get(i, 'campo'),
                'auto_tipo': col_tipos.get(i, 'text'),
                'is_campo': auto_mapping.get(i) == 'campo',
            }
            for i, h in enumerate(headers)
        ]

        return render(request, 'contacts/importar_preview.html', {
            'header_data': header_data,
            'headers': headers,
            'preview_rows': rows[:10],
            'total_rows': len(rows),
            'filename': file.name,
            'roles': [
                ('nombre', 'Nombre *'), ('telefono', 'Teléfono *'),
                ('email', 'Email'), ('grupo', 'Grupo'),
                ('notas', 'Notas'), ('campo', 'Campo personalizado'),
                ('ignorar', 'Ignorar'),
            ],
            'tipos': TIPOS_CAMPO,
        })

    def _handle_confirm(self, request):
        tmp_path = request.session.get('import_tmp')
        ext = request.session.get('import_ext', 'xlsx')
        headers = request.session.get('import_headers', [])

        if not tmp_path or not os.path.exists(tmp_path):
            messages.error(request, 'Sesión expirada. Volvé a subir el archivo.')
            return redirect('contacts:importar')

        col_roles = {}
        col_tipos = {}
        for i in range(len(headers)):
            role = request.POST.get(f'rol_{i}', 'ignorar')
            col_roles[str(i)] = role
            if role == 'campo':
                col_tipos[str(i)] = request.POST.get(f'tipo_{i}', 'text')

        update_existing = request.POST.get('update_existing') == 'on'
        agregar_prefijo_ar = request.POST.get('agregar_prefijo_ar') == 'on'

        with open(tmp_path, 'rb') as f:
            headers_full, rows_full, error = parse_file(f, ext)

        try:
            os.unlink(tmp_path)
        except Exception:
            pass
        request.session.pop('import_tmp', None)

        if error:
            messages.error(request, error)
            return redirect('contacts:importar')

        created, updated, skipped, errors = import_from_rows(
            headers_full, rows_full, col_roles, col_tipos, update_existing,
            agregar_prefijo_ar=agregar_prefijo_ar,
        )

        msg = f'Importación completa: {created} creados, {updated} actualizados, {skipped} omitidos.'
        if errors:
            msg += f' {len(errors)} error(es).'
            for e in errors[:5]:
                messages.warning(request, e)
        messages.success(request, msg)
        return redirect('contacts:list')


# ──────────────────────────────────────────────
# APIs JSON
# ──────────────────────────────────────────────

class ContactoBuscarAPIView(LoginRequiredMixin, View):
    def get(self, request):
        q = request.GET.get('q', '').strip()
        if len(q) < 1:
            return JsonResponse({'contactos': []})
        qs = Contacto.objects.filter(
            Q(nombre__icontains=q) | Q(telefono__icontains=q)
        )[:10]
        return JsonResponse({'contactos': [
            {'id': c.pk, 'nombre': c.nombre, 'telefono': c.telefono}
            for c in qs
        ]})


class CamposParaGrupoAPIView(LoginRequiredMixin, View):
    def get(self, request):
        grupo = request.GET.get('grupo', '')
        campos = _campos_para_grupo(grupo).values('pk', 'etiqueta', 'tipo', 'grupo')
        return JsonResponse({'campos': list(campos)})


# ──────────────────────────────────────────────
# Grupos
# ──────────────────────────────────────────────

class GruposListView(LoginRequiredMixin, View):
    def get(self, request):
        from django.db.models import Count
        grupos = (
            Contacto.objects
            .exclude(grupo='')
            .values('grupo')
            .annotate(total=Count('pk'))
            .order_by('grupo')
        )
        sin_grupo = Contacto.objects.filter(grupo='').count()
        total_contactos = Contacto.objects.count()
        return render(request, 'contacts/grupos_list.html', {
            'grupos': list(grupos),
            'sin_grupo': sin_grupo,
            'total_contactos': total_contactos,
        })


class GrupoAsignarView(LoginRequiredMixin, View):
    template_name = 'contacts/grupo_asignar.html'

    def _ctx(self):
        campos = list(
            CampoPersonalizado.objects.filter(activo=True)
            .order_by('orden', 'etiqueta')
            .values('pk', 'etiqueta', 'tipo', 'grupo')
        )
        grupos = list(
            Contacto.objects.values_list('grupo', flat=True)
            .distinct().exclude(grupo='').order_by('grupo')
        )
        return {
            'campos_disponibles': campos,
            'grupos': grupos,
            'fecha_opciones': FECHA_OPCIONES,
            'operadores_por_tipo': json.dumps(OPERADORES_POR_TIPO),
        }

    def get(self, request):
        ctx = self._ctx()
        ctx['grupo_inicial'] = request.GET.get('grupo', '')
        return render(request, self.template_name, ctx)

    def post(self, request):
        grupo_destino = request.POST.get('grupo_destino', '').strip()
        accion = request.POST.get('accion_grupo', 'asignar')

        if not grupo_destino and accion == 'asignar':
            messages.error(request, 'Ingresá un nombre de grupo.')
            ctx = self._ctx()
            ctx['post_data'] = request.POST
            return render(request, self.template_name, ctx)

        filtros = parse_filtros_from_post(request.POST)
        qs = apply_filters(filtros)

        if accion == 'quitar':
            # Remove from group only the contacts that currently have that group
            count = qs.filter(grupo=grupo_destino).update(grupo='')
            messages.success(request, f'{count} contacto{"s" if count != 1 else ""} removido{"s" if count != 1 else ""} del grupo "{grupo_destino}".')
        else:
            count = qs.count()
            if not count:
                messages.warning(request, 'No hay contactos que coincidan con los filtros.')
                ctx = self._ctx()
                ctx['post_data'] = request.POST
                return render(request, self.template_name, ctx)
            qs.update(grupo=grupo_destino)
            messages.success(request, f'{count} contacto{"s" if count != 1 else ""} asignado{"s" if count != 1 else ""} al grupo "{grupo_destino}".')

        return redirect('contacts:grupos')


class GrupoEliminarView(LoginRequiredMixin, View):
    """Remove a group label from ALL its contacts (sets grupo='')."""
    def post(self, request):
        grupo = request.POST.get('grupo', '').strip()
        if not grupo:
            return redirect('contacts:grupos')
        count = Contacto.objects.filter(grupo=grupo).update(grupo='')
        messages.success(request, f'Grupo "{grupo}" eliminado ({count} contactos sin grupo).')
        return redirect('contacts:grupos')


# ──────────────────────────────────────────────
# Helper
# ──────────────────────────────────────────────

def _auto_link_conversacion(telefono, contacto, nombre):
    try:
        from apps.whatsapp.models import Conversacion
        Conversacion.objects.filter(
            telefono=telefono, contacto__isnull=True
        ).update(contacto=contacto, nombre_contacto=nombre)
    except Exception:
        pass
